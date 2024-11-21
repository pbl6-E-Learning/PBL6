import time
import openpyxl
import os
from selenium.webdriver.common.by import By
from selenium import webdriver
from PIL import Image
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage
from base import setup_driver

def create_result_excel():
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Search Results"
  ws.append(["Level", "Kết quả tìm kiếm", "Screenshot"])
  return wb, ws

def save_screenshot_to_excel(driver, level, ws, row):
  screenshot = driver.get_screenshot_as_png()
  image = Image.open(BytesIO(screenshot))
  image_path = f"{level}.png"
  image.save(image_path)
  img = ExcelImage(image_path)
  img.width = 400
  img.height = 300
  ws.add_image(img, f"C{row}")

def create_results_directory(directory_path):
  if not os.path.exists(directory_path):
    os.makedirs(directory_path)

def test_filters_and_save_screenshot():
  results_dir = 'test/Create_course/results'
  create_results_directory(results_dir)
  wb, ws = create_result_excel()
  driver = setup_driver()

  try:
    for idx, level in enumerate(['N1', 'N2', 'N3', 'N4', 'N5'], start=2):
      time.sleep(7)
      filter_button = driver.find_element(By.XPATH, "//button[.//span[contains(text(), 'Filter by Level')]]")
      filter_button.click()
      time.sleep(10)
      level_option = driver.find_element(By.XPATH, f"//div[@role='option' and .//span[text()='{level}']]")
      level_option.click()
      search_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Search')]")
      search_button.click()
      time.sleep(2)
      results = driver.find_elements(By.TAG_NAME, "tr")
      result_text = f"{len(results)-1} courses found" if len(results)-1 > 0 else "No courses found"
      ws[f"A{idx}"] = level
      ws[f"B{idx}"] = result_text
      save_screenshot_to_excel(driver, level, ws, row=idx)
      driver.refresh()

    result_file_path = os.path.join(results_dir, 'filter_results_with_screenshots.xlsx')
    wb.save(result_file_path)
    print(f"File Excel đã được tạo và lưu tại: {result_file_path}")

  finally:
    driver.quit()
