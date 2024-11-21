import time
import openpyxl
import os
from selenium.webdriver.common.by import By
from selenium import webdriver
from PIL import Image
from io import BytesIO
from openpyxl.drawing.image import Image as ExcelImage
from base import setup_driver

def read_search_data_from_excel(file_path):
  wb = openpyxl.load_workbook(file_path)
  ws = wb.active
  search_keywords = []
  for row in ws.iter_rows(min_row=2, max_col=1, values_only=True):
    search_keywords.append(row[0])
  return search_keywords

def save_screenshot_to_excel(driver, search_term, ws, row):
    screenshot = driver.get_screenshot_as_png()
    image = Image.open(BytesIO(screenshot))
    buffer = BytesIO()
    image.save(buffer, format="PNG")
    buffer.seek(0)
    img = ExcelImage(buffer)
    img.width = 400
    img.height = 300
    ws.add_image(img, f"B{row}")

def create_result_excel(search_data):
  wb = openpyxl.Workbook()
  ws = wb.active
  ws.title = "Search Results"
  ws.append(["Tên", "Kết quả tìm kiếm"])
  for index, term in enumerate(search_data, start=2):
    ws.append([term, f"Kết quả cho {term}"])
  return wb, ws

def create_results_directory(directory_path):
  if not os.path.exists(directory_path):
    os.makedirs(directory_path)

def test_search_and_save_screenshot():
    search_data = read_search_data_from_excel('test/Create_course/data/search_data.xlsx')
    results_dir = 'test/Create_course/results'
    create_results_directory(results_dir)
    wb, ws = create_result_excel(search_data)
    driver = setup_driver()
    try:
        for index, term in enumerate(search_data, start=2):
            time.sleep(7)
            search_input = driver.find_element(By.XPATH, "//input[@placeholder='Search Course']")
            search_input.clear()
            search_input.send_keys(term)
            search_button = driver.find_element(By.XPATH, "//button[contains(text(), 'Search')]")
            search_button.click()
            time.sleep(7)
            save_screenshot_to_excel(driver, term, ws, index)
            td_elements = driver.find_elements(By.TAG_NAME, "tr")
            course_count = len(td_elements) - 1
            result_text = f"{course_count} courses found" if course_count > 0 else "No courses found"
            ws[f"B{index}"] = result_text
        result_file_path = os.path.join(results_dir, 'search_results_with_screenshots.xlsx')
        wb.save(result_file_path)
        print(f"File Excel đã được tạo và lưu tại: {result_file_path}")
    finally:
        driver.quit()
